
import os
from datetime import datetime
from os import path
from random import choice, randint

import openpyxl
from xlsxwriter import Workbook

from cogestione import db as database

def delete_corso(id):
    try:
        id = int(id)
    except ValueError:
        print("not int")

    db = database.get_db()

    corso = db.session.scalar(db.select(database.corso).where(database.corso.id == id))
    organizzazioni = db.session.scalars(db.select(database.organizza).join(database.corso).where(database.corso.id == id)).all()

    if corso is None:
        print("not ok")

    for org in organizzazioni:
        db.session.delete(org)
    db.session.delete(corso)
    db.session.commit()

    print("ok")

def create_corso(titolo, descrizione, organizzatori, note, fascia):
    fascia = int(fascia)

    db = database.get_db()

    fascia1 = fascia == 1
    fascia2 = fascia == 2
    fascia3 = fascia == 3
    fascia4 = fascia == 4
    fascia5 = fascia == 5

    fasce = [fascia1, fascia2, fascia3, fascia4, fascia5]

    fasce = [i+1 for i, el in enumerate(fasce) if el]

    if len(fasce) == 0:
        print("no fascia")
        return

    organizzatori_str = organizzatori
    organizzatori = [x for x in organizzatori.split(";") if x not in ["", None]]
    organizzatori = list(set(organizzatori))

    for f in fasce:
        corso = database.corso(titolo, descrizione, 30, 0, "non ancora assegnata", f, organizzatori_str, note)
        db.session.add(corso)

        cnt_ref = 0
        for org_email in organizzatori:
            user_org = db.session.scalar(db.select(database.user).filter_by(email = org_email))
            if user_org is None:
                continue

            prv_org = db.session.scalars(db.select(database.organizza).join(database.user).join(database.corso).where(database.user.id == user_org.id, database.corso.fascia == f)).all()

            if len(prv_org) != 0:
                continue

            organizza = database.organizza(user_org.id, corso.id)
            cnt_ref += 1
            db.session.add(organizza)

        # if cnt_ref == 0:
        #     db.session.delete(corso)


    db.session.commit()

    print("ok")



def fissa_aula(id_corso, id_aula):
    db = database.get_db()
    id_corso = int(id_corso)
    id_aula = int(id_aula)

    corso = db.session.scalar(db.select(database.corso).where(database.corso.id == id_corso))

    corso.aula_id = id_aula
    corso.aula_fissata = True

    db.session.commit()

    print("ok")


def add_aula(nome, posti):
    db = database.get_db()

    db.session.add(database.aula(
        nome = nome,
        posti_totali = posti,
    ))
    db.session.commit()

    print("ok")


def carica_aule(path):

    if not os.path.exists(path):
        print("path does not exists")
        return

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    if ws is None:
        return

    db = database.get_db()
    for value in ws.iter_rows(min_row=2):
        nome = f"Edificio {value[0].value}, piano {value[1].value}, aula {value[2].value}"
        posti = int(str(value[3].value))
        db.session.add(database.aula(
            nome = nome,
            posti_totali = posti,
        ))
        db.session.commit()



def carica_utenti(path):

    if not os.path.exists(path):
        print("path does not exists")
        return

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    if ws is None:
        return

    db = database.get_db()
    for value in ws.iter_rows(min_row=2):
        db.session.add(database.user(
            email = str(value[6].value),
            nome = str(value[1].value) if value[1].value is not None else "",
            cognome = str(value[0].value),
            classe = str(value[5].value),
            password = ""
        ))
        db.session.commit()



def carica_prof(path):

    if not os.path.exists(path):
        print("path does not exists")
        return

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    if ws is None:
        return

    db = database.get_db()
    for value in ws.iter_rows(min_row=2):
        db.session.add(database.user(
            email = str(value[1].value),
            nome = "",
            cognome = str(value[0].value),
            classe = "",
            password = ""
        ))
        db.session.commit()


def create_xlsx_file(file_path: str, headers: dict, items: list):
    with Workbook(file_path) as workbook:
        worksheet = workbook.add_worksheet()
        worksheet.write_row(row=0, col=0, data=headers.values())
        header_keys = list(headers.keys())
        for index, item in enumerate(items):
            row = map(lambda field_id: item.get(field_id, ''), header_keys)
            worksheet.write_row(row=index + 1, col=0, data=row)

def students_sheet():
    date = str(datetime.now().time())
    date = date[:date.find('.')]

    dir_path = path.join(os.path.dirname(__file__), "output-data")
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)

    file_name = path.join(os.path.dirname(__file__), "output-data", "dati-finali" + ".xlsx")
    # file_name = "dati-" + date + ".xlsx"

    headers = {
        "nome" : "Nome",
        "cognome" : "Cognome",
        "classe" : "Classe",
        "titolo_corso" : "Titolo Corso",
        "aula" : "Aula",
        "fascia" : "Fascia",
        "organizza" : "Organizza",
    }

    db = database.get_db()

    items = []
    iscrizioni = db.session.scalars(db.select(database.iscrizione)).all()
    organizzazioni = db.session.scalars(db.select(database.organizza)).all()


    for iscrizione in iscrizioni:
        if iscrizione.user is None:
            continue
        items.append({
            "nome" : iscrizione.user.nome,
            "cognome" : iscrizione.user.cognome,
            "classe" : iscrizione.user.classe,
            "titolo_corso" : iscrizione.corso.titolo,
            "aula" : iscrizione.corso.aula,
            "fascia" : iscrizione.corso.fascia,
            "organizza" : ""
        })

    for iscrizione in organizzazioni:
        if iscrizione.user is None:
            continue
        items.append({
            "nome" : iscrizione.user.nome,
            "cognome" : iscrizione.user.cognome,
            "classe" : iscrizione.user.classe,
            "titolo_corso" : iscrizione.corso.titolo,
            "aula" : iscrizione.corso.aula,
            "fascia" : iscrizione.corso.fascia,
            "organizza" : "Organizzatore"
        })


    create_xlsx_file(file_name, headers, items)

    print("Done!")


def randomizzato(lista : list, func):
    if len(lista) == 0:
        raise BaseException("List is empty")

    totale = sum(list(map(func, lista)))
    n = randint(0, totale+1)

    contatore = 0
    for x in lista:
        contatore += func(x)
        if contatore >= n:
            return x
    return choice(lista)

def riassegna_posti_occupati():

    db = database.get_db()

    corsi = db.session.scalars(db.select(database.corso)).all()

    for corso in corsi:
        cnt = len(db.session.scalars(db.select(database.iscrizione).where(database.iscrizione.corso_id == corso.id)).all())
        corso.posti_occupati = cnt

    db.session.commit()

def forza_iscrizioni():

    db = database.get_db()

    utenti = db.session.scalars(db.select(database.user)).all()

    for i in range(1, 6):
        corsi = list(db.session.scalars(db.select(database.corso).where(database.corso.fascia == i, database.corso.posti_occupati != database.corso.posti_totali)).all())
        print(f"fascia {i}:")

        for user in utenti:

            if user.classe == "": # docente
                continue

            isc = db.session.scalar(db.select(database.iscrizione).join(database.corso).join(database.user).where(database.corso.fascia == i, database.user.id == user.id))
            org = db.session.scalar(db.select(database.organizza).join(database.corso).join(database.user).where(database.corso.fascia == i, database.user.id == user.id))

            if isc is None and org is None:
                corso = randomizzato(corsi, lambda x : x.posti_totali - x.posti_occupati)

                db.session.add(database.iscrizione(user_id = user.id, corso_id = corso.id))
                corso.posti_occupati += 1

                if corso.posti_occupati >= corso.posti_totali:
                    corsi.remove(corso)

            print(f"user {user.id} / {i}")

        db.session.commit()

    print("Fatto!")


def fix_emails(userid, email):
    db = database.get_db()
    userid = int(userid)

    sel = db.select(database.user).where(database.user.id == userid)
    u = db.session.scalar(sel)
    if u is None:
        print("not ok")
        return

    u.email = email
    db.session.commit()
    print("ok")
