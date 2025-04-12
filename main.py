
from flask import Flask
from flask import redirect, url_for
from flask import render_template
from flask import request, jsonify, json
from flask import session
from flask import flash
from flask import make_response, Response
from datetime import timedelta, datetime
from flask_sqlalchemy import SQLAlchemy
from helper import *
import database
from database import db, app
from io import StringIO
from contextlib import redirect_stdout
import re
from os import path
from werkzeug.security import check_password_hash, generate_password_hash
from xlsxwriter import Workbook
import openpyxl
from random import randint, choice
import requests
import os
import dotenv
import sys

DB_NAME = "database.db"
# app = Flask(__name__)
# app.secret_key = "secret key"
# app.permanent_session_lifetime = timedelta(days=7)
# app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_NAME}'


@app.route("/")
def home():
    return render_template("homepage.html")


@app.route("/admin", methods=["GET", "POST"])
@admin_access
def execute():
    if request.method == "POST":
        
        headers = dict(request.headers)
        dati = json.loads(request.data)
        res = ""
        
        f = StringIO()
        with redirect_stdout(f):
            try:
                exec(dati["comando"])
                res = f.getvalue().strip()
            except Exception as e:
                res = str(e)
        
        # headers["Result"] = str(res)
        return jsonify({"res" : res})
        return Response([b"good"], headers=headers)
    elif request.method == "GET":
        return render_template("admin.html")
    return "idk"

def fix_emails(nome, cognome, classe, email):
    sel = db.select(database.user).where(database.user.nome.contains(nome.strip().upper()), database.user.cognome.contains(cognome.strip().upper()), database.user.classe.contains(sanitize_classe(classe)))
    
    if len(list(db.session.scalars(sel).all())) > 1:
        print("C'è più di un utente con questi robi")
        return
    
    db.session.scalars(sel).first().email = email
    db.session.commit()

def solve_emails(path):
    if not os.path.exists(path):
        print(path)
        print("path does not exists")
        return
          
    
    wb = openpyxl.load_workbook(path)
    ws = wb.active


    for value in ws.iter_rows(min_row=2):
        nome = value[1].value if value[1].value != None else ""
        nome = nome.lower()
        cognome = value[0].value.lower()
        classe = value[2].value
        email = value[3].value
        email = email.strip()
        if not email.endswith("@liceolussana.eu") and classe != "docente":
            email = f"{cognome}.{nome}.studente@liceolussana.eu"
            fix_emails(nome, cognome, classe, email)


def carica_corsi(path):
    
    print(path)

    if not os.path.exists(path):
        print("path does not exists")
        return
    
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    # values = [ws.cell(row=1,column=i).value for i in range(1,ws.max_column+1)]
    
    
    FASCIA = 0
    for value in ws.iter_rows(min_row=2):
        titolo = value[1].value
        referenti = value[2].value
        descrizione = value[3].value
        aula = value[4].value
        capienza = value[5].value
        real_referenti = value[6].value
        
        if real_referenti is None:
            real_referenti = ""
        
        if not all([titolo, referenti, descrizione, aula, capienza]):
            if titolo is not None and titolo.startswith("FASCIA"):
                FASCIA = int(titolo[7])
            continue

        if str(titolo).strip().lower().count("nome") != 0 and str(aula).strip().lower().count("aula") != 0:
            # print(f"--------------------------------------------------------------------------------  CORSI FASCIA {FASCIA}  --------------------------------------------------------------------------------")
            continue

        try:
            new_capienza = ""
            for c in str(capienza):
                if c.isnumeric():
                    new_capienza += c
            capienza = int(new_capienza)
        except:
            print(f"error reading capienza: {capienza} - {new_capienza}")
            print(titolo, referenti, descrizione, aula)
            continue
        
        
        new_corso = database.corso(
            titolo = titolo,
            descrizione = descrizione,
            posti_totali = capienza,
            posti_occupati = 0,
            aula = aula,
            fascia = FASCIA,
            organizzatori_str = referenti
        )
        db.session.add(new_corso)
        db.session.commit()
        
        for x in real_referenti.split(";"):
            user = db.session.scalars(db.select(database.user).where(database.user.email == x)).first()
            if user is None:
                print(x)
                continue
                # raise "vaffanculo"
            db.session.add(database.organizza(
                utente = user.id,
                corso = new_corso.id
            ))
        
        # print(f"{FASCIA} --- {titolo}:\n referenti: {referenti} \n descrizione: {descrizione} \n aula: {aula} \n capienza: {capienza} \n\n\n")
                
    # pass

def carica_utenti(path):

    if not os.path.exists(path):
        print("path does not exists")
        return
    
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    
    for value in ws.iter_rows(min_row=2):
        db.session.add(database.user(
            email = value[3].value,
            nome = value[1].value if value[1].value != None else "",
            cognome = value[0].value,
            classe = value[2].value,
            password = ""
        ))
        db.session.commit()
        
        # anagrafica.append((
        #     str(value[0].value).strip().lower(),
        #     value[1].value.strip().lower() if value[1].value is not None else "",
        #     sanitize_classe(str(value[2].value).strip().lower()),
        #     str(value[3].value).strip().lower()
        # ))

# statistiche

@app.route("/report/", methods=["GET"])
def report():
    dati = {}
    info_classe = {}
    
    lista_classi = []
    path = os.path.join(os.path.dirname(__file__), "input_data", "lista_classi.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for value in ws.iter_rows(min_row=1):
        if value[0].value is not None:
            lista_classi.append(value[0].value)
    
    
    sel = db.select(database.iscrizione)
    iscrizioni = db.session.scalars(sel).all()

    sel = db.select(database.organizza)
    organizzazioni = db.session.scalars(sel).all()

    sel = db.select(database.user)
    utenti = db.session.scalars(sel).all()
    
    for classe in lista_classi:
        dati.setdefault(classe, dict())
    
    for iscrizione in iscrizioni:
        if iscrizione.userref is None or iscrizione.corsoref is None:
            continue
        if iscrizione.corsoref.id == 44:
            dati.setdefault(iscrizione.userref.classe, dict())
            dati[iscrizione.userref.classe].setdefault(iscrizione.corsoref.fascia+1, 0)
            dati[iscrizione.userref.classe][iscrizione.corsoref.fascia+1] += 1
            
        dati.setdefault(iscrizione.userref.classe, dict())
        dati[iscrizione.userref.classe].setdefault(iscrizione.corsoref.fascia, 0)
        dati[iscrizione.userref.classe][iscrizione.corsoref.fascia] += 1

    for organizza in organizzazioni:
        if organizza.userref is None:
            continue
        if organizza.corsoref.id == 44:
            dati.setdefault(organizza.userref.classe, dict())
            dati[organizza.userref.classe].setdefault(organizza.corsoref.fascia+1, 0)
            dati[organizza.userref.classe][organizza.corsoref.fascia+1] += 1
            
        dati.setdefault(organizza.userref.classe, dict())
        dati[organizza.userref.classe].setdefault(organizza.corsoref.fascia, 0)
        dati[organizza.userref.classe][organizza.corsoref.fascia] += 1
    
    for utente in utenti:
        info_classe.setdefault(utente.classe, 0)
        info_classe[utente.classe] += 1
    
    for classe in lista_classi:
        dati.setdefault(classe, dict())
        d = dati[classe]
        # print(classe, d)
        for i in range(1, 6):
            dati[classe].setdefault(i, 0)
            dati[classe][i] = round(dati[classe][i] / info_classe[classe] * 100)
    
    
    
    return render_template("report.html", dati=dati)

def create_xlsx_file(file_path: str, headers: dict, items: list):
    with Workbook(file_path) as workbook:
        worksheet = workbook.add_worksheet()
        worksheet.write_row(row=0, col=0, data=headers.values())
        header_keys = list(headers.keys())
        for index, item in enumerate(items):
            row = map(lambda field_id: item.get(field_id, ''), header_keys)
            worksheet.write_row(row=index + 1, col=0, data=row)

def fix_dati():
    organizzazioni = db.session.scalars(db.select(database.organizza)).all()
    iscrizioni = db.session.scalars(db.select(database.iscrizione)).all()
    
    s = {}
    for org in organizzazioni:
        t = (org.userref.id, org.corsoref.id)
        s[t] = 1

    count = 0
    for iscrizione in iscrizioni:
        t = (iscrizione.userref.id, iscrizione.corsoref.id)
        s.setdefault(t, 0)
        if s[t] == 1:
            count += 1
            db.session.delete(iscrizione)
        
        s[t] = 1

    print(f"double iscrizioni: {count}")
    db.session.commit()

def randomizzato(lista : list, func):
    if len(lista) == 0:
        raise BaseException("List is empty")

    totale = sum(list(map(func, lista)))
    n = randint(0, totale+1)
    
    contatore = 0
    for x in lista:
        contatore += func(x)
        if contatore >= n:
            return x
    return choice(lista)

def forza_iscrizioni():

    organizzazioni = db.session.scalars(db.select(database.organizza)).all()
    utenti = db.session.scalars(db.select(database.user)).all()

    for i in range(1, 6):
        iscrizioni = db.session.scalars(db.select(database.iscrizione)).all()
        corsi = list(db.session.scalars(db.select(database.corso).where(database.corso.fascia == i, database.corso.posti_occupati != database.corso.posti_totali)).all())
        if i == 2:
            corsi.append(db.session.scalars(db.select(database.corso).where(database.corso.id == 44)).first())
        print(f"fascia {i}:")
        for user in utenti:
            if user.classe == 'docente':
                continue
            le_sue_iscrizioni = [x for x in iscrizioni if x.userref.id == user.id]
            le_sue_organizzazioni = [x for x in organizzazioni if x.userref.id == user.id]
            ids = [x.corsoref.id for x in le_sue_iscrizioni]
            ids2 = [x.corsoref.id for x in le_sue_organizzazioni]
            ids += ids2
            fasce = [x.corsoref.fascia for x in le_sue_iscrizioni]
            fasce2 = [x.corsoref.fascia for x in le_sue_organizzazioni]
            fasce += fasce2
            if 44 in ids:
                fasce.append(2)

            if i not in fasce:
                if i == 1 and 2 not in fasce:
                    corso = corsi[-1]
                else:
                    corso = randomizzato(corsi, lambda x : x.posti_totali - x.posti_occupati)
                while (corso.id == 44 and 2 in fasce) or (corso.id == 44 and 1 in fasce):
                    # print(len(corsi), " - ".join(map(lambda x : str(x.posti_occupati), corsi)))
                    # print(corso)
                    if len(corsi) <= 1 and corso.id == 44:
                        db.session.commit()
                        corsi = list(db.session.scalars(db.select(database.corso).where(database.corso.fascia == i, database.corso.posti_occupati != database.corso.posti_totali)).all())
                        print(list(map(lambda x : str(x.posti_totali - x.posti_occupati), corsi)))
                        
                        print("merdaccia")
                        exit(0)
                        raise BaseException("sto cazzo di capitol di merda porcaccia")
                    corso = randomizzato(corsi, lambda x : x.posti_totali - x.posti_occupati)
                
                db.session.add(database.iscrizione(utente=user.id, corso=corso.id))
                corso.posti_occupati += 1
                
                if corso.posti_occupati == corso.posti_totali:
                    corsi.remove(corso)
            print(f"user {user.id} / {i}")

        db.session.commit()
    return "Fatto!"

def dati_per_corso_per_fascia():

    dir_path = path.join('.', "output-data", "corsi")
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)

    
    corsi = db.session.scalars(db.select(database.corso)).all()
    corsi = list(corsi)
    corsi = sorted(corsi, key=lambda x : x.fascia)
    
    iscrizioni = db.session.scalars(db.select(database.iscrizione)).all()
    
    fascia = 0
    
    for corso in corsi:
        if corso.fascia > fascia:
            fascia = corso.fascia
            dir_path = path.join('.', "output-data", "corsi", f"fascia {fascia}")
            if not os.path.exists(dir_path):
                os.makedirs(dir_path)
        

        file_name = path.join('.', "output-data", "corsi", f"fascia {fascia}", f"{corso.id}.xlsx")
        headers = {
            "nome" : "Nome",
            "cognome" : "Cognome",
            "classe" : "Classe"
        }
        items = []
        for x in iscrizioni:
            if x.corsoref.id != corso.id:
                continue
            items.append({
                "nome" : x.userref.nome,
                "cognome" : x.userref.cognome,
                "classe" : x.userref.classe
            })
        create_xlsx_file(file_name, headers, items)
    
def corsi_finali():
    
    date = str(datetime.now().time())
    date = date[:date.find('.')]

    dir_path = path.join('.', "output-data")
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    file_name = path.join('.', "output-data", "corsi-finali-" + date + ".xlsx")
    
    headers = {
        "fascia" : "",
        "id" : "ID",
        "nome" : "Nome corso",
        "aula" : "Aula",
        "persone" : "Persone"
    }
    
    items = []
    corsi = db.session.scalars(db.select(database.corso)).all()
    corsi = list(corsi)
    corsi = sorted(corsi, key=lambda x : x.fascia)
    
    iscrizioni = db.session.scalars(db.select(database.iscrizione)).all()
    
    fascia = 0
    
    for corso in corsi:
        if corso.fascia > fascia:
            fascia = corso.fascia
            items.append({
                "fascia" : "",
                "id" : "",
                "nome" : "",
                "aula" : "",
                "persone" : ""
            })
            items.append({
                "fascia" : f"Fascia {fascia}",
                "id" : "",
                "nome" : "",
                "aula" : "",
                "persone" : ""
            })
        posti_occupati = len([x for x in iscrizioni if x.corsoref.id == corso.id])
        items.append({
            "fascia" : "",
            "id" : f"{corso.id}",
            "nome" : f"{corso.titolo}",
            "aula" : f"{corso.aula}",
            "persone" : f"{posti_occupati}"
        })



    create_xlsx_file(file_name, headers, items)
    
    return "fatto"

def assegna_sorveglianza():
    
    path = os.path.join('.', "input_data", "SORVEGLIANZA.xlsx")

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    
    corsi_sorv = [0 for _ in range(5)]
    for i in range(5):
        if db.session.scalars(db.select(database.corso).where(database.corso.titolo == f"sorveglianza {i+1}")).first() is None:
            corso = database.corso(
                titolo=f"sorveglianza {i+1}",
                fascia=i+1
            )
            db.session.add(corso)
            db.session.commit()
        corsi_sorv[i] = db.session.scalars(db.select(database.corso).where(database.corso.titolo == f"sorveglianza {i+1}")).first().id
    

    for value in ws.iter_rows(min_row=2):
        for i in range(0, 5):
            if value[2*i+1].value is None:
                continue
            email = value[2*i+1].value.strip()

            user = db.session.scalars(db.select(database.user).where(database.user.email == email)).first()
            # iscrizioni = db.session.scalars(db.select(database.iscrizione).join(database.user).where(database.user.email == email)).all()
            # organizzazioni = db.session.scalars(db.select(database.organizza).join(database.user).where(database.user.email == email)).all()
            # print(user)
            # print(iscrizioni, organizzazioni)
            # print([x.corsoref for x in iscrizioni])
            if user is None:
                print(email)
                continue
            # print(user.id)
            for x in db.session.scalars(db.select(database.iscrizione).join(database.user).join(database.corso).where(database.user.id == user.id, database.corso.fascia == i+1)):
                db.session.delete(x)
            db.session.commit()
            # fasce_iscrizioni = [x.corsoref.fascia for x in iscrizioni if x.corsoref is not None]
            # fasce_organizzazioni = [x.corsoref.fascia for x in organizzazioni if x.corsoref is not None]
            
            # if i+1 in fasce_iscrizioni or i+1 in fasce_organizzazioni:
            #     # print(f"user {user.id} is already busy in fascia {i}")
            #     continue
            
            db.session.add(database.iscrizione(corso=corsi_sorv[i], utente=user.id))
        db.session.commit()
        
    return "fatto!"

def save_data():
    date = str(datetime.now().time())
    date = date[:date.find('.')]
    
    dir_path = path.join('.', "output-data")
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    file_name = path.join('.', "output-data", "dati-" + date + ".xlsx")
    # file_name = "dati-" + date + ".xlsx"
    
    headers = {
        "id" : "Id",
        "nome" : "Nome",
        "cognome" : "Cognome",
        "classe" : "Classe",
        "titolo_corso" : "Titolo Corso",
        "aula" : "Aula",
        "fascia" : "Fascia",
    }
    
    items = []
    iscrizioni = db.session.scalars(db.select(database.iscrizione)).all()
    
    for iscrizione in iscrizioni:
        if iscrizione.userref is None:
            continue
        items.append({
            "id" : iscrizione.id,
            "nome" : iscrizione.userref.nome,
            "cognome" : iscrizione.userref.cognome,
            "classe" : iscrizione.userref.classe,
            "titolo_corso" : iscrizione.corsoref.titolo,
            "aula" : iscrizione.corsoref.aula,
            "fascia" : iscrizione.corsoref.fascia
        })
    
    create_xlsx_file(file_name, headers, items)
    
    return "Done!"

# azioni

@app.route("/lista-corsi/<n_fascia>", methods=["GET"])
def lista_corsi(n_fascia):
    corsi = db.session.execute(db.select(database.corso).filter_by(fascia=n_fascia)).all()
    
    # corsi = database.corso.query.filter_by(id=n_fascia).all()
    corsi = list(map(lambda x : x[0], corsi))
    return render_template("lista-corsi.html", fascia=int(n_fascia), corsi=corsi, dim=len(corsi))

@app.route("/lista-corsi/", methods=["GET"])
def lista_corsi_help():
    return redirect(request.url + "1")

@app.route("/iscrizione/", methods=["POST"])
@login_required
def iscrizione():
    flash("Le iscrizioni sono chiuse")
    return Response([b"error"], 400)

    dati = json.loads(request.data)
    id_corso = dati["id_corso"]
    
    user = db.session.execute(db.select(database.user).filter_by(email=session["email"])).first()[0]
    corso = db.session.execute(db.select(database.corso).filter_by(id=id_corso)).first()[0]
    
    if corso is None:
        flash("Il corso non esiste", 'error')
        return Response([b"error"], 400)
    
    if corso.posti_occupati == corso.posti_totali:
        flash("Il corso è pieno!", 'error')
        return Response([b"error"], 400)
    
    # iscrizioni = db.session.execute(db.select(database.iscrizione).filter_by())
    # print(user)
    # print(user[0].iscrizioni)
    # iscrizioni = user.iscrizioni

    sel = db.select(database.iscrizione).join(database.user).where(database.user.email == session["email"])
    iscrizioni = db.session.scalars(sel).all()
    
    sel = db.select(database.organizza).join(database.user).where(database.user.email == session["email"])
    organizzazioni = db.session.scalars(sel).all()
    # iscrizioni = db.session.scalars(db.select(database.iscrizione).where(database.iscrizione.userref.email == session["email"])).all()
    
    #(db.select(database.user, database.iscrizione, database.corso).filter_by(email=session["email"])).all()

    # print(iscrizioni)
    
    for iscrizione in iscrizioni:
        if corso.fascia <= 2 and iscrizione.corsoref.id == 44:
            flash("Sei già iscritto a un corso nella prima fascia che dura 2 fasce. Puoi annullare l'iscrizione dal tuo profilo", 'error')
            return Response([b"error"], 400)
            
        if corso.id == 44:
            if iscrizione.corsoref.fascia <= 2:
                flash("Sei già iscritto a un corso per questa fascia o per la successiva. Questo corso dura 2 fasce. Puoi annullare l'iscrizione dal tuo profilo", 'error')
                return Response([b"error"], 400)

        if iscrizione.corsoref.fascia == corso.fascia:
            flash("Sei già iscritto a un corso per questa fascia. Puoi annullare l'iscrizione dal tuo profilo", 'error')
            return Response([b"error"], 400)

    for organizza in organizzazioni:
        if corso.fascia <= 2 and organizza.corsoref.id == 44:
            flash("Sei già organizzatore di un corso per questa fascia", 'error')
            return Response([b"error"], 400)
        
        if corso.id == 44:
            if organizza.corsoref.fascia <= 2:
                flash("Sei già organizzatore di un corso per questa fascia o per la successiva. Questo corso dura 2 fasce", 'error')
                return Response([b"error"], 400)

        if organizza.corsoref.fascia == corso.fascia:
            flash("Sei già organizzatore di un corso per questa fascia", 'error')
            return Response([b"error"], 400)
    
    
    # print("iscritto a", id_corso)
    db.session.add(database.iscrizione(utente=user.id, corso=corso.id))
    corso.posti_occupati += 1
    db.session.commit()
    
    flash("Iscritto con successo", 'success')
    return Response([b"good"], 200)

@app.route("/annulla-iscrizione/", methods=["POST"])
@login_required
def annulla_iscrizione():
    flash("Le iscrizioni sono chiuse e non puoi più modificarle")
    return Response([b"error"], 400)

    dati = json.loads(request.data)
    id_corso = dati["id_corso"]
    
    corso = db.session.execute(db.select(database.corso).filter_by(id=id_corso)).first()[0]
    sel = db.select(database.iscrizione).join(database.user).join(database.corso).where(database.user.email == session["email"], database.corso.id == id_corso)
    iscrizione = db.session.scalars(sel).first()
    
    if iscrizione is None:
        return Response([b"bad"], 404)
    
    db.session.delete(iscrizione)
    corso.posti_occupati -= 1
    db.session.commit()
    
    return jsonify({})

@app.route("/corso/<id_corso>", methods=["GET"])
def info_corso(id_corso):
    corso = database.corso.query.filter_by(id=id_corso).first()
    referenti = " - ".join(list(map(lambda x : f"{x.userref.nome} {x.userref.cognome} {x.userref.classe}", corso.organizzatori)))
    return render_template("corso.html", corso=corso, referenti=referenti)

@app.route("/appello/<id_corso>", methods=["GET", "POST"])
@login_required
def appello(id_corso):
    id_corso = int(id_corso)
    
    corso = db.session.scalars(db.select(database.corso).where(database.corso.id == id_corso)).first()

    if corso is None:
        return redirect(url_for("home"))
    
    organizza = db.session.scalars(db.select(database.organizza).join(database.user).where(database.user.email == session["email"], database.organizza.corso == id_corso)).first()
    
    if organizza is None:
        return redirect(url_for("home"))
    
    if request.method == "GET":
        iscrizioni = db.session.scalars(db.select(database.iscrizione).join(database.corso).where(database.corso.id == id_corso)).all()
        organizzazioni = db.session.scalars(db.select(database.organizza).join(database.corso).where(database.corso.id == id_corso)).all()
        
        presenza = db.session.scalars(db.select(database.presenza).join(database.corso).where(database.corso.id == id_corso)).all()

        persone = []
        
        for x in (iscrizioni + organizzazioni):
            assegnato = False
            valore = None
            for v in presenza:
                if v.userref.id == x.userref.id:
                    assegnato = True
                    valore = v.presente
                    break
            
            persone.append({
                "id" : x.userref.id,
                "name" : f"{x.userref.nome} {x.userref.cognome}",
                "assegnato" : assegnato,
                "valore" : valore
            })

        return render_template("appello.html", persone=persone, corso=corso)
    
    # POST request
    dati = json.loads(request.data)
    
    iscrizioni = db.session.scalars(db.select(database.iscrizione).join(database.corso).where(database.corso.id == id_corso)).all()
    organizzazioni = db.session.scalars(db.select(database.organizza).join(database.corso).where(database.corso.id == id_corso)).all()
    
    presenza = db.session.scalars(db.select(database.presenza).join(database.corso).where(database.corso.id == id_corso)).all()
    
    id_presenti = [x.userref.id for x in (iscrizioni + organizzazioni)]

    
    for chiave in dati:
        key = int(chiave)
        val = dati[chiave]
        if key not in id_presenti:
            continue
        for x in presenza:
            if x.userref.id == key:
                x.presente = val
                break
        else:
            db.session.add(database.presenza(utente=key, corso=id_corso, presente=val))
    
    db.session.commit()
    flash("Presenze aggiornate con successo", 'success')
    return Response([b"good"], 200)

# profilo

@app.route("/profile/")
@login_required
def profile():
    
    user = db.session.execute(db.select(database.user).filter_by(email=session["email"])).first()[0]
    
    sel = db.select(database.iscrizione).join(database.user).where(database.user.email == session["email"])
    iscrizioni = db.session.scalars(sel).all()
    
    sel = db.select(database.organizza).join(database.user).where(database.user.email == session["email"])
    organizzazioni = db.session.scalars(sel).all()
    
    corsi = [dict() for _ in range(6)]
    
    for iscrizione in iscrizioni:
        if iscrizione.corsoref.id == 44:
            d = corsi[iscrizione.corsoref.fascia+1]
            d["id"] = iscrizione.corsoref.id
            d["titolo"] = iscrizione.corsoref.titolo
            d["posti"] = f"{iscrizione.corsoref.posti_occupati} / {iscrizione.corsoref.posti_totali}"
            d["organizzatori"] = " - ".join(list(map(lambda x : f"{x.userref.nome} {x.userref.cognome} {x.userref.classe}", iscrizione.corsoref.organizzatori)))
            d["aula"] = iscrizione.corsoref.aula
            d["organizzato"] = False
        
        d = corsi[iscrizione.corsoref.fascia]
        d["id"] = iscrizione.corsoref.id
        d["titolo"] = iscrizione.corsoref.titolo
        d["posti"] = f"{iscrizione.corsoref.posti_occupati} / {iscrizione.corsoref.posti_totali}"
        d["organizzatori"] = " - ".join(list(map(lambda x : f"{x.userref.nome} {x.userref.cognome} {x.userref.classe}", iscrizione.corsoref.organizzatori)))
        d["aula"] = iscrizione.corsoref.aula
        d["organizzato"] = False
        # d["annulla iscrizione"] = f"<button onclick=\"annulla_iscrizione({iscrizione.corsoref.id})\"> Annulla </button>"
    
    for organizza in organizzazioni:
        if organizza.corsoref.id == 44:
            d = corsi[organizza.corsoref.fascia+1]
            d["id"] = organizza.corsoref.id
            d["titolo"] = organizza.corsoref.titolo
            d["posti"] = f"{organizza.corsoref.posti_occupati} / {organizza.corsoref.posti_totali}"
            d["organizzatori"] = " - ".join(list(map(lambda x : f"{x.userref.nome} {x.userref.cognome} {x.userref.classe}", organizza.corsoref.organizzatori)))
            d["aula"] = organizza.corsoref.aula
            d["organizzato"] = True

        d = corsi[organizza.corsoref.fascia]
        d["id"] = organizza.corsoref.id
        d["titolo"] = organizza.corsoref.titolo
        d["posti"] = f"{organizza.corsoref.posti_occupati} / {organizza.corsoref.posti_totali}"
        d["organizzatori"] = " - ".join(list(map(lambda x : f"{x.userref.nome} {x.userref.cognome} {x.userref.classe}", organizza.corsoref.organizzatori)))
        d["aula"] = organizza.corsoref.aula
        d["organizzato"] = True
    
    
    return render_template("profile.html", corsi=corsi, utente=user, orari_fasce=orari_fasce)

# accesso al profilo
@app.route("/login/", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html")
    
    email = request.form.get("email").strip()
    password = request.form.get("password").strip()
    
    user = db.session.execute(db.select(database.user).filter_by(email=email)).first()
    
    if user is None:
        flash("Email o password errati", 'error')
        return render_template("login.html")
    
    if not check_password_hash(user[0].password, password):
    # if user[0].password != password:
        flash("Email o password errati", 'error')
        return render_template("login.html")

    session.permanent = True
    session["email"] = email
    session["logged"] = True
    
    flash("Login effettuato con successo", 'success')
    return redirect(url_for("lista_corsi", n_fascia=1))

def send_email(to_email, subject, content):
    dotenv.load_dotenv()
    res = requests.post(
        "https://api.eu.mailgun.net/v3/cogestione-lussana.eu/messages",
        auth=("api", os.getenv('API_KEY', 'API_KEY')),
        data={"from": "Iscrizioni Cogestione Lussana <iscrizioni@cogestione-lussana.eu>",
            "to": to_email,
            "subject": subject,
            "text": content
            })
    
    # print(to_email, res)

def send_auth_verification_email(email):
    testo = f"""Questo è il tuo codice di verififca:\n\t{session["auth_code"]}\n\nInseriscilo nella pagina di registrazione!"""
    
    send_email(email,
               "Codice di verifica cogestione",
               testo)

def send_pwd_reset_email(email):
    testo = f"""Questo è il tuo codice di verifica:\n\t{session["auth_code"]}\n\nInseriscilo nella pagina di reset della password!"""
    
    send_email(email, "Codice di verifica cogestione", testo)


@app.route("/verification/", methods=["GET", "POST"])
def verification():
    if session.get("auth_code", -1) == -1 or session.get("auth_age", -1) == -1:
        return redirect(url_for("home"))
    
    if (datetime.now() - datetime.fromisoformat(session["auth_age"])).seconds >= 600:
        session.clear()
    
    if session.get("auth_code", -1) == -1 or session.get("auth_age", -1) == -1:
        return redirect(url_for("home"))
    
    
    if request.method == "GET":
        return render_template("verification.html")
    
    code = request.form.get("verification-code")
    
    if code != session["auth_code"]:
        session.clear()
        flash("Il codice di controllo è sbagliato", 'error')
        return redirect(url_for("register"))
    
    # user = session["user"]
    email = session["tmp_email"]
    # nome = session["tmp_nome"]
    # cognome = session["tmp_cognome"]
    # classe = session["tmp_classe"]
    password = session["tmp_password"]
    session.clear()
    
    user = db.session.execute(db.select(database.user).filter_by(email=email)).first()
    user[0].password = password
    
    # user = database.user(email=email, nome=nome, cognome=cognome, classe=classe, password=password)
    # db.session.add(user)
    db.session.commit()
    
    flash("Registrato con successo", 'success')
    return redirect(url_for("login"))

@app.route("/register/", methods=["GET", "POST"])
def register():
    if request.method == "GET":
        return render_template("register.html")
    
    failed = False
    
    # nome = request.form.get("name")
    # cognome = request.form.get("lastname")
    email = request.form.get("email").strip()
    # classe = request.form.get("classe")
    password1 = request.form.get("password1").strip()
    password2 = request.form.get("password2").strip()
    
    if re.match(EMAIL_REGEX, email) is None:
        flash("Email non valida", 'error')
        failed = True

    if not email.endswith("@liceolussana.eu"):
        flash("Devi usare l'email istituzionale (...@liceolussana.eu)")
        failed = True
    
    # classe = sanitize_classe(classe)
    
    # if classe is None or False:
    #     flash("La classe non è valida", 'error')
    #     failed = True
    
    # more checks
    
    user = db.session.execute(db.select(database.user).filter_by(email=email)).first()
    
    if user is None:
        flash("L'email non è esistente. Se credi ci sia stato un errore, contatta i rappresentanti", "error")
        failed = True
        return render_template("register.html")
    
    # print(user)
    if user[0].password != "":
        flash("Email è già in uso", 'error')
    
    if password1 != password2:
        flash("Le password non corrispondono", 'error')
        failed = True

    if failed:
        return render_template("register.html")
    
    session.clear()
    session["auth_code"] = "".join([str(randint(0, 9)) for _ in range(6)])
    # session["auth_code"] = "000000"
    session["auth_age"] = datetime.now().isoformat()
    
    flash(f"Ti abbiamo inviato una mail di verifica su {email}")
    
    send_auth_verification_email(email)
    
    password = generate_password_hash(password1)

    session["tmp_email"] = email
    # session["tmp_nome"] = nome
    # session["tmp_cognome"] = cognome
    # session["tmp_classe"] = classe
    session["tmp_password"] = password

    return redirect(url_for("verification"))

@app.route("/verification-reset-pwd/", methods=["GET", "POST"])
def verification_reset_pwd():
    if session.get("auth_code", -1) == -1 or session.get("auth_age", -1) == -1:
        return redirect(url_for("home"))
    
    if (datetime.now() - datetime.fromisoformat(session["auth_age"])).seconds >= 600:
        session.clear()
    
    if session.get("auth_code", -1) == -1 or session.get("auth_age", -1) == -1:
        return redirect(url_for("home"))
    
    
    if request.method == "GET":
        return render_template("verification-reset-pwd.html")
    
    code = request.form.get("verification-code")
    
    if code != session["auth_code"]:
        session.clear()
        flash("Il codice di controllo è sbagliato", 'error')
        return redirect(url_for("reset_password"))
    
    email = session["tmp_email"]
    password = session["tmp_password"]
    session.clear()
    
    user = db.session.execute(db.select(database.user).filter_by(email=email)).first()[0]
    user.password = password
    db.session.commit()
    session.clear()
    
    flash("Password modificata con successo", 'success')
    return redirect(url_for("login"))

@app.route("/reset-password/", methods=["GET", "POST"])
def reset_password():
    if request.method == "GET":
        return render_template("reset-pwd.html")
    
    failed = False
    
    email = request.form.get("email")
    password1 = request.form.get("password1")
    password2 = request.form.get("password2")
    
    # more checks
    
    user = db.session.execute(db.select(database.user).filter_by(email=email)).first()
    
    if user is None:
        flash("L'email non è valida", "error")
        failed = True
    
    if password1 != password2:
        flash("Le password non corrispondono", 'error')
        failed = True

    if failed:
        return render_template("reset-pwd.html")
    
    session.clear()
    session["auth_code"] = "".join([str(randint(0, 9)) for _ in range(6)])
    session["auth_age"] = datetime.now().isoformat()
    
    flash(f"Ti abbiamo inviato una mail di verifica su {email}")
    
    send_pwd_reset_email(email)
    
    password = generate_password_hash(password1)
    
    session["tmp_email"] = email
    session["tmp_password"] = password

    return redirect(url_for("verification_reset_pwd"))

@app.route("/logout/")
@login_required
def logout():
    session.clear()
    return redirect(url_for("home"))


if __name__ == "__main__":
    dotenv.load_dotenv()
    
    with app.app_context():
        # db.drop_all()
        # db.init_app(app)
        # db.create_all()
        if database.init_db(app):
            print("carica")
            dirPath = os.path.dirname(__file__)
            carica_utenti(os.path.join(dirPath, "input_data", "Anagrafica.xlsx"))
            carica_corsi(os.path.join(dirPath, "input_data", "Corsi-fixed.xlsx"))
        
        # debug = False on production
        app.run(debug=True)
